import numpy as np   #pip install numpy
import pytesseract   #pip install pytesseract
import cv2   #pip install opencv-python

from PIL import Image

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract'

#imegen = cv2.imread(r"C:\Users\User\Videos\Captures\Prueba1.png",0)
#image = Image.open(r"C:\Users\User\Videos\Captures\Prueba1.png")
""" imagen = cv2.imread('img_00.jpeg')
grices = cv2.cvtColor(imagen, cv2.COLOR_BGR2GRAY)

canny = cv2.Canny(grices, 10, 150)

canny = cv2.dilate(canny, None, iterations=1)



texto = pytesseract.image_to_string(image=image)

print(texto)


cv2.imshow('Prueba', imegen)
cv2.waitKey(1000 * 4)
cv2.destroyAllWindows()



Capture = cv2.VideoCapture(0)

while(Capture.isOpened()):
    ret,imagen = Capture.read()
    if ret is True:
        cv2.imshow('video', imagen)
        if cv2.waitKey(1) & 0xFF == ord('t'):
            break

Capture.release()
cv2.destroyAllWindows()
"""


import pytesseract
from PIL import Image
import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv
import os

load_dotenv()

# Establece la ruta al tesseract
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# Cargar la imagen y extraer texto con pytesseract
def imagen_a_texto(ruta_de_la_imagen):
    imagen = Image.open(ruta_de_la_imagen)
    texto = pytesseract.image_to_string(imagen)
    return texto

# Procesar el texto extraído para transformarlo en lista
def procesar_texto(texto):
    data = [linea for linea in texto.strip().split("\n") if linea != '']
    return data

# Guardar o actualizar el archivo Excel
def save_to_excel(data, excel_path):
    try:
        workbook = load_workbook(excel_path)
        sheet = workbook.active
        start_row = sheet.max_row + 1  # Añadir nuevos datos al final
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        workbook.save(excel_path)
    except FileNotFoundError:
        # Si el archivo no existe, creamos uno nuevo
        df = pd.DataFrame(data)
        df.to_excel(excel_path, index=False, header=False)

# Ruta de la imagen y el archivo Excel
#image_path = 'ruta_a_tu_imagen.jpg'
#excel_path = 'ruta_a_tu_archivo.xlsx'

# Procesamiento de OCR y almacenamiento en Excel
#text = image_to_text(image_path)
#data = process_text(text)
#save_to_excel(data, excel_path)

#Extrae la ruta de la imagen del .env
RUTA_IMAGEN = os.getenv('RUTA_IMAGEN')

texto_extraido = imagen_a_texto(fr'{RUTA_IMAGEN}')

#Trasforma el texto en lista para procesar el Dataframe
lista_texto = procesar_texto(texto_extraido)

#Extre los posiciones necesarias para el dataframe
index_Year = lista_texto.index('Year')+1
index_Category = lista_texto.index('Category')+1
index_Product = lista_texto.index('Product')+1
index_Sales = lista_texto.index('Sales')+1
index_Rating = lista_texto.index('Rating')+1

#Crea el diccionario
info_a_excel = {'Year' : lista_texto[index_Year:index_Category-1],
                'Category' : lista_texto[index_Category:index_Product-1],
                'Product' : lista_texto[index_Product:index_Sales-1],
                'Sales' : lista_texto[index_Sales:index_Rating-1],
                'Rating' : lista_texto[index_Rating: ]}


#Revisa cual es la lista más larga
max_len = max(len(col) for col in info_a_excel.values())

#Rellena las listas que no tienen la cantidad necesaria con Nan
for key in info_a_excel:
    info_a_excel[key].extend([np.nan] * (max_len - len(info_a_excel[key])))

#Transforma el diccionario a Dataframe
info_a_excel = pd.DataFrame(info_a_excel)

#Extrae la ruta donde se creara el excel del .env
RUTA_EXCEL = os.getenv('RUTA_EXCEL')

name = (Fr'{RUTA_EXCEL}')

#Convierte el dataframe a excel
with pd.ExcelWriter(name) as ExcelWriter:
    info_a_excel.to_excel(excel_writer=ExcelWriter, index=False)

""" print(a)

print(a.index('Year'))
sublista1 = a[a.index('Year')+1:a.index('Category')]
print(sublista1)
print(len(sublista1))
print(a.index('Category'))
sublista2 = a[a.index('Category')+1:a.index('Product')]
print(sublista2)
print(len(sublista2))
print(a.index('Product'))
sublista3 = a[a.index('Product')+1:a.index('Sales')]
print(sublista3)
print(len(sublista3))
print(a.index('Sales'))
sublista4 = a[a.index('Sales')+1:a.index('Rating')]
print(sublista4)
print(len(sublista4))
print(a.index('Rating'))
sublista5 = a[a.index('Rating')+1:]
print(sublista5)

print(len(sublista5))"""